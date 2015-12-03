# vision.reader
# Reads the Excel spreadsheet of our finances and emits account values.
#
# Author:   Benjamin Bengfort <bengfort@cs.umd.edu>
# Created:  Wed Dec 02 17:59:13 2015 -0500
#
# Copyright (C) 2015 Bengfort.com
# For license information, see LICENSE.txt
#
# ID: vision.reader.py [] benjamin@bengfort.com $

"""
Reads the Excel spreadsheet of our finances and emits account values.
"""

##########################################################################
## Imports
##########################################################################

import re
import decimal

from datetime import datetime
from operator import add, sub
from openpyxl import load_workbook

##########################################################################
## Module Constants
##########################################################################

## Worksheet Parse Phases
HEADER       = 0
ACCOUNTS     = 1
TRANSACTIONS = 2

## CENTS
CENTS = decimal.Decimal('0.01')

##########################################################################
## Simple function expressions
##########################################################################

xlfn = re.compile('^=Table\d+\[.+\](.+)$', re.I)
opfn = re.compile('[\-\+][A-Z\d]+')
ops  = {
    '+': add,
    '-': sub,
}

def monetize(value):
    return decimal.Decimal(value).quantize(CENTS, decimal.ROUND_05UP)

##########################################################################
## Account Data Structure
##########################################################################

class FinanceSheet(object):
    """
    Data structure for holding financial sheet information.
    """

    def __init__(self, date):
        self.date = date
        self.accounts = []
        self.transactions = []

    def add_account(self, account, ws):
        """
        Handles the acount and adds it to the accounts list.
        """
        if not isinstance(account['Ending Balance'], float):
            fn = xlfn.match(account['Ending Balance'])
            if fn is not None:
                mods = opfn.findall(fn.groups()[0])

                bal = account['Beginning Balance']
                for mod in mods:
                    op   = ops[mod[0]]
                    cell = mod[1:]

                    bal = op(bal, ws[cell].value)

                account['Ending Balance'] = bal

        # Convert beginning and ending balance to decimals
        account['Beginning Balance'] = monetize(account['Beginning Balance'])
        account['Ending Balance'] = monetize(account['Ending Balance'])
        self.accounts.append(account)

    def add_transaction(self, transaction):
        """
        Handles the transaction and adds it to the transactions list.
        """
        transaction['Date'] = transaction['Date'].date()
        transaction['Amount'] = monetize(transaction['Amount'])
        self.transactions.append(transaction)

##########################################################################
## Financial Reader
##########################################################################

class SpreadsheetReader(object):
    """
    Reads the worksheets of the Finances.xlsx files created by Bengfort.com
    """

    def __init__(self, path):
        self.path  = path
        self.wb    = load_workbook(filename=path, read_only=True)

    @property
    def sheets(self):
        return self.wb.get_sheet_names()

    def rows(self, sheet):
        """
        Yields the rows for the given sheet.
        """
        ws = self.wb[sheet]

        for row in ws.rows:
            row = [cell.value for cell in row][:4]
            row = filter(lambda a: a is not None, row)
            if row: yield row

    def finances(self, month):
        """
        Parses and returns a finances data structure.
        This is kind of a shit show, lots of branching, but hopefully works.
        """
        ws     = self.wb[month]
        phase  = HEADER
        sheet  = None
        actype = None
        fields = None

        for row in self.rows(month):

            if phase == HEADER:
                if len(row) == 1 and isinstance(row[0], datetime):
                    sheet = FinanceSheet(row[0].date())

                elif row[0].lower() == "accounts":
                    if sheet is None:
                        raise ValueError("No finances sheet date detected!")
                    phase = ACCOUNTS

                else:
                    raise ValueError("Unknown HEADER row: {!r}".format(row))

            elif phase == ACCOUNTS:
                if row[0].lower() == "transactions":
                    phase = TRANSACTIONS

                elif len(row) == 1:
                    actype = row[0]

                elif len(row) == 3:
                    # Skip totals
                    continue

                else:
                    if row[0].lower() == "bank":
                        fields = row
                    else:
                        account = dict(zip(fields, row))
                        account['Account Type'] = actype
                        sheet.add_account(account, ws)

            elif phase == TRANSACTIONS:

                if isinstance(row[0], datetime):
                    sheet.add_transaction(dict(zip(fields, row)))
                elif row[0].lower() == "date":
                    fields = row
                else:
                    raise ValueError("Unknown TRANSACTION row: {!r}".format(row))

            else:
                raise Exception("Unknown phase, bad handling!")

        return sheet
